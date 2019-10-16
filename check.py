from __future__ import annotations
import requests
import sys
from openpyxl import Workbook
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border,Alignment
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
import urllib3
import socket
import dns.resolver
from dns.rdatatype import CNAME, A
from sty import fg, rs
from urllib.parse import urlparse
from urllib.parse import urljoin
from checker import DnsCategory, DnsCheck
from checker.DnsCheck import DnsCheck
from checker.DnsCategory import DnsCategory
from typing import Dict, List, Tuple

if len(sys.argv) is not 3:
    print('''{}Usage: python {} [FILE] [MY_DOMAIN] [FORMAT]{}
    [FILE] zone file
    [MY_DOMAIN] treat this domain as 1st party    
    '''.format(fg.red, sys.argv[0], fg.rs))

    exit(-1)

TIMEOUT = 1.0
DEBUG = True
MY_DOMAIN = sys.argv[2]


def check(proto, domain):
    if DEBUG:
        print('[HTTP] Checking {}'.format('{}://{}'.format(proto, domain)))
    checks = list()
    http_check = check_url('{}://{}'.format(proto, domain))
    checks.append(http_check)
    # follow redirects, step by step
    depth = 0
    while http_check['code'] not in ['SSL', 'LOOP', 'ERR', 'TIME'] and http_check['code'] >= 300 and http_check['code'] < 400:
        if DEBUG:
            print('{}Following redirect to {}.{}'.format(fg.red, http_check['url'], fg.rs))
        depth += 1
        if depth > 7:
            print('{}Error: check resulted in too many redirects "{}"{}'.format(fg.red, http_check['url'], fg.rs))
            break
        # handling relative URLs
        next_url = http_check['redirect']
        if urlparse(next_url).scheme == '':
            if DEBUG:
                print('check({}, {}) fixing relative URL {}'.format(
                    proto, domain, next_url), end='')
            if len(checks) < 2:
                next_url = urljoin(proto + '://' + domain, next_url)
            else:
                prev_url = urlparse(checks[-2]['url'])
                next_url = urljoin(prev_url.scheme + '://' + prev_url.netloc, next_url)
            if DEBUG:
                print(' to {}'.format(next_url))

        http_check = check_url(next_url)
        checks.append(http_check)
    return checks


def check_url(url):
    try:
        res = requests.get(url, allow_redirects=False, timeout=TIMEOUT)
        if res.status_code >= 300 and res.status_code < 400:
            return {'code': res.status_code, 'url': url, 'redirect': res.headers['Location']}
        else:
            return {'code': res.status_code, 'url': url}
    except requests.exceptions.Timeout:
        return {'code': 'TIME', 'url': f'timeout after {TIMEOUT} seconds'}
    except requests.exceptions.SSLError:

        # print(e.args[0].reason)
        # print(type(e.args[0].reason).__name__)
        # print('vars')
        # print(vars(e.args[0]))
        return {'code': 'SSL'}  # {e.args[0].reason}'}
    except requests.exceptions.TooManyRedirects:
        return {'code': 'LOOP'}
    except:
        # print(sys.exc_info()[0], end='')
        return {'code': 'ERR', 'url': f'error {sys.exc_info()[0].__name__}'}


def format_code(checks):
    out = ""
    for check in checks:
        code = check['code']
        if type(code) is str:
            out += ' [{}{}{}]'.format(fg.red, code, fg.rs)
        elif 200 == code:
            out += ' [{}{}{}]'.format(fg.green, code, fg.rs)
        elif code >= 300 and code < 400:
            if not MY_DOMAIN in check['url']:
                color = fg.red
            else:
                color = fg.green
            target = ' -> {}{}{}'.format(color, check['url'], fg.rs)
            out += ' [{}{}{}]'.format(fg.li_black, code, fg.rs) + target
        elif code >= 400 and code < 500:
            out += ' [{}{}{}]'.format(fg.yellow, code, fg.rs)
        else:
            out += ' [{}{}{}]'.format(fg.red, code, fg.rs)
    return out


def read_zonefile() -> List[DnsCheck]:
    '''
        read zonefile from disk, removing duplicate entries
    '''
    entries = list()
    seen = set()
    filename = sys.argv[1]
    if DEBUG:
        print('{}Reading Zone file from {}.{}'.format(fg.green, filename, fg.rs))

    with open(filename) as f:
        for line in f.readlines():
            s = list(filter(lambda x: x != '', line.split()))

            if not len(s) == 5 or s[3] not in ('CNAME', 'A'):
                continue

            # strip the '.' at end]
            domain = s[0][:-1]
            typ = s[3]
            if domain not in seen:
                seen.add(domain)
                entries.append(DnsCheck(domain, typ))

            if DEBUG:
                print(f'testing {DnsCheck(domain, typ)}')
    return entries


def check_dns(entry: DnsCheck) -> DnsCheck:
    '''
    check DNS entries

    -> 1st party
    -> 3rd party
    -> DNS Error
    '''

    try:
        answers = dns.resolver.query(entry.domain, entry.typ, lifetime=TIMEOUT)
        if DEBUG:
            print('{}[DNS]: query [{}]{} {}'.format(fg.yellow, entry.typ, entry.domain, fg.rs), end='')

        for answer in answers:

            if DEBUG:
                print('{} answer {}. {}'.format(fg.yellow, answer, fg.rs), end='')

            if answer.rdtype == CNAME:
                target: str = answer.target.to_text(omit_final_dot=True)
                entry.status = 'ok'
                if DEBUG:
                    print('{} is aliased to {}'.format(entry.domain, target))

                if target.endswith(MY_DOMAIN):
                    entry.category = DnsCategory.first_party
                else:
                    entry.category = DnsCategory.third_party

                # result is still a CNAME --> recursively check them, too
                if DEBUG:
                    print('{}[DNS]: following CNAME chain for {}.{}'.format(fg.green, DnsCheck(target, 'CNAME'), fg.rs))
                for query_type in ['CNAME', 'A', 'AAAA']:
                    sub_check = check_dns(DnsCheck(target, query_type))
                    if sub_check.status == 'ok':
                        break
                entry.sub_check = sub_check

            if answer.rdtype == A:
                if DEBUG:
                    print('[DNS]: {} resolved to address {}'.format(entry.domain, answer.address))
                entry.status = 'ok'
                entry.category = DnsCategory.a_record

            if not answer.rdtype in [A, CNAME]:
                print('[DNS]: Unhandled rdataset.rdtype={} for entry={}'.format(answer.rdtype, entry))

    except dns.resolver.NXDOMAIN:
        if DEBUG:
            print("{}[DNS]: No such domain {}.{}".format(fg.yellow, entry.domain, fg.rs))
        entry.status = 'nok'
        entry.category = DnsCategory.dns_error
    except dns.resolver.NoAnswer:
        if DEBUG:
            print("{}[DNS]: NoAnswer [{}]{}.{}".format(fg.yellow, entry.typ, entry.domain, fg.rs))
        if entry.domain.endswith('acm-validations.aws'):
            # expected NoAnswer for AWS' DNS based certificate validation
            entry.status = 'ok'
            entry.category = DnsCategory.dns_validation
        else:
            entry.status = 'nok'
            entry.category = DnsCategory.dns_error
    except dns.resolver.NoNameservers:
        if DEBUG:
            print("{}[DNS]: NoNameservers {}.{}".format(fg.yellow, entry.domain, fg.rs))
        entry.status = 'nok'
        entry.category = DnsCategory.dns_error
    except dns.exception.Timeout:
        print('{}[DNS]: Timeout {}.{}'.format(fg.yellow, entry.domain, fg.rs))
        entry.status = 'nok'
        entry.category = DnsCategory.dns_error

    return entry


def is_check_not_decent(check):
    error_codes = ['SSL', 'LOOP', 'ERR', 'TIME']

    is_http_error = check['http_check'][-1]['code'] in error_codes
    is_https_error = check['https_check'][-1]['code'] in error_codes
    is_dns_error = check['classification'].effective_category().value >= DnsCategory.dns_error.value
    is_dns_cert_validation = check['classification'].effective_category() == DnsCategory.dns_validation

    return not is_dns_cert_validation and (is_http_error or is_https_error or is_dns_error)


def csv_2(fileName: str, output, only_errors: bool):
    wb = Workbook()
    ws = wb.active
    ws.append(['Record name', 'Type', 'Category', 'HTTP', 'HTTPS', 'HTTP trace', 'HTTPS trace', 'DNS trace'])

    for i in list(filter(lambda x: x['classification'].status != 'ok', output)):
        ws.append([
            i['classification'].domain,
            i['classification'].typ,
            i['classification'].effective_category().name,
            '',
            '',
            '',
            ''
        ])
    
    if only_errors:
        filtered_list = list(filter(lambda x: x['classification'].status == 'ok' and is_check_not_decent(x), output))
    else:
        filtered_list = list(filter(lambda x: x['classification'].status == 'ok', output))
    
    for i in filtered_list:

        http_trace = [path['url'] if 'url' in path else path['code']
                      for path in i['http_check']]
        https_trace = [path['url'] if 'url' in path else path['code']
                       for path in i['https_check']]

        ws.append([
            i['classification'].domain,
            i['classification'].typ,
            i['classification'].effective_category().name,
            i['http_check'][-1]['code'],
            i['https_check'][-1]['code'],
            "\n-> ".join(http_trace),
            "\n-> ".join(https_trace),
            i['classification'].dns_trace()
        ])

    def as_text(value):
        if value is None:
            return ""
        return str(value)

    # auto-size columns
    for column_cells in ws.columns:
        for cell in column_cells:            
            ws[cell.coordinate].alignment = Alignment(wrapText=True)
            ws[cell.coordinate].alignment = Alignment(wrapText=True)
            ws[cell.coordinate].alignment = Alignment(wrapText=True)            
        clumn_max_length = len(max( max(as_text(cell.value).split('\n') for cell in column_cells) ))
        column_letter = (get_column_letter(column_cells[0].column))
        if clumn_max_length > 0:
            ws.column_dimensions[column_letter].width = clumn_max_length + 5

    red_fill = PatternFill(bgColor="FFC7CE")
    green_fill = PatternFill(bgColor="00C700")
    yellow_fill = PatternFill(bgColor="FED51A")

    ws.conditional_formatting.add('D2:E' + str(ws.max_row), CellIsRule(operator='between', formula=['200', '299'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('D2:E' + str(ws.max_row), CellIsRule(operator='between', formula=['400', '499'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('D2:E' + str(ws.max_row), CellIsRule(operator='between', formula=['500', '599'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('D2:E' + str(ws.max_row), CellIsRule(operator='between', formula=['"A"', '"Z"'], stopIfTrue=True, fill=red_fill))

    ws.conditional_formatting.add('C2:C' + str(ws.max_row), CellIsRule(operator='==', formula=['"first_party"'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('C2:C' + str(ws.max_row), CellIsRule(operator='==', formula=['"a_record"'], stopIfTrue=True, fill=green_fill))
    ws.conditional_formatting.add('C2:C' + str(ws.max_row), CellIsRule(operator='==', formula=['"third_party"'], stopIfTrue=True, fill=yellow_fill))
    ws.conditional_formatting.add('C2:C' + str(ws.max_row), CellIsRule(operator='==', formula=['"dns_error"'], stopIfTrue=True, fill=red_fill))
    ws.conditional_formatting.add('C2:C' + str(ws.max_row), CellIsRule(operator='==', formula=['"dns_validation"'], stopIfTrue=True, fill=green_fill))
    
    wb.save(fileName)
    print('CSV was saved to {}'.format(fileName))


def csv(output):
    csv_2("all.xlsx", output, only_errors=False)
    csv_2("errors.xlsx", output, only_errors=True)


if __name__ == '__main__':
    # if False:
    if DEBUG:
        entries: List[DnsCheck] = list()
        # this is no longer registered
        entries.append(DnsCheck('sterne-1.welt.de', 'A'))
        # all ok
        entries.append(DnsCheck('www.welt.de', 'A'))
        # broken CNAME
        entries.append(DnsCheck('aktion.welt.de', 'CNAME'))
        # redirect
        entries.append(DnsCheck('bmw.welt.de', 'CNAME'))
        entries.append(DnsCheck('amp.welt.de', 'CNAME'))
        entries.append(DnsCheck('www.beste.welt.de', 'CNAME'))
        # acm validation
        entries.append(DnsCheck('_269a588eb6ae0cd44533c9d13f48808f.amp.welt.de', 'CNAME'))
    else:
        entries: List[DnsCheck] = read_zonefile()

    output = list()
    cnt = 0

    print("Found {} items to be checked".format(len(entries)))

    for entry in entries:
        cnt += 1
        print('Processing {}/{}                   '.format(cnt, len(entries)), end='\r')

        classification = check_dns(entry)

        if DEBUG:
            print('{}[DNS]: result {} {}\n\n'.format(fg.green, classification, fg.rs))

        if DEBUG and cnt == 20:
            break

        if classification.status == 'ok':

            domain = entry.domain

            http_checks = check('http', domain)
            https_checks = check('https', domain)

            output.append({
                'classification': classification,
                'http_check': http_checks,
                'https_check': https_checks,
                'pretty': 'HTTP {}\nHTTPS{}'.format(format_code(http_checks), format_code(https_checks))
            })
        else:
            output.append({
                'classification': classification
            })
    print("\n")    
    csv(output)
